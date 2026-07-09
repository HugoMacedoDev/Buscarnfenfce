#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SIGA SEFAZ-CE - Coletor de XMLs de NF-e via Certificado A1
Este script realiza o download dos XMLs completos de NF-e da Receita Federal.
- Para notas de DESTINATÁRIO (Entradas): Utiliza o método de consulta por NSU (sem limites e muito rápido).
- Para notas de EMISSOR (Saídas): Utiliza o método de consulta por Chave, com limitador de 20 consultas por hora para evitar bloqueios.
"""

import os
import sys
import json
import time
import base64
import re
import csv
import gzip
import tempfile
import xml.etree.ElementTree as ET
import urllib3
import requests

# Garante compatibilidade de caracteres especiais/emojis no console Windows
if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

# Desativa avisos de SSL/TLS
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

try:
    import openpyxl
except ImportError:
    print("\n❌ A biblioteca 'openpyxl' é necessária para ler as planilhas Excel.")
    print("Execute o seguinte comando no terminal para instalá-la:")
    print("pip install openpyxl")
    sys.exit(1)

try:
    from cryptography.hazmat.primitives.serialization import pkcs12
    from cryptography.hazmat.primitives import serialization
except ImportError:
    print("\n❌ A biblioteca 'cryptography' é necessária para processar certificados A1.")
    print("Execute o seguinte comando no terminal para instalá-la:")
    print("pip install cryptography")
    sys.exit(1)

EXCEL_FILE = "empresas cod.xlsx"

# Histórico global de timestamps de consultas por chave para o controle de limite (20/hora)
query_timestamps = []

def wait_for_rate_limit():
    """Garante que não ultrapassaremos 20 requisições por chave por hora por certificado."""
    global query_timestamps
    now = time.time()
    # Filtra apenas os timestamps dos últimos 3600 segundos (1 hora)
    query_timestamps = [t for t in query_timestamps if now - t < 3600]
    
    if len(query_timestamps) >= 20:
        # Calcula quanto tempo falta para a requisição mais antiga expirar da janela de 1 hora
        wait_time = 3600 - (now - query_timestamps[0]) + 5  # Margem de segurança de 5s
        print(f"\n⚠️ Limite de 20 consultas/hora atingido. Aguardando {int(wait_time)} segundos (cerca de {int(wait_time/60)} min) para liberar nova consulta...")
        time.sleep(wait_time)
        # Atualiza a lista pós-espera
        now = time.time()
        query_timestamps = [t for t in query_timestamps if now - t < 3600]
        
    query_timestamps.append(time.time())

def normalize_name(name):
    """Normaliza o nome da empresa para cruzamento flexível."""
    if not name:
        return ""
    name = name.upper()
    name = re.sub(r'[^A-Z0-9\s]', '', name)
    endings = [
        r'\bLTDA\b', r'\bME\b', r'\bEPP\b', r'\bEIRELI\b', r'\bSA\b', r'\bS\b',
        r'\bLIMITADA\b', r'\bMICROEMPRESA\b', r'\bMICRO\b', r'\bEMPRESA\b'
    ]
    for ending in endings:
        name = re.sub(ending, '', name)
    return " ".join(name.split())

def sanitize_filename(filename):
    """Remove caracteres inválidos no Windows."""
    return re.sub(r'[\\/*?:"<>|]', "", filename)

def load_excel_mapping(filepath):
    """Lê a planilha Excel com os códigos internos e nomes das empresas."""
    if not os.path.exists(filepath):
        print(f"❌ Planilha Excel '{filepath}' não encontrada no diretório atual.")
        sys.exit(1)
        
    print(f"📖 Carregando planilha '{filepath}'...")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active
    
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print("❌ A planilha Excel está vazia.")
        sys.exit(1)
        
    headers = [str(h).strip().lower() for h in rows[0] if h is not None]
    code_idx = 0
    name_idx = 1
    
    for idx, h in enumerate(headers):
        if "cod" in h:
            code_idx = idx
        elif "nome" in h or "empresa" in h or "raz" in h:
            name_idx = idx
            
    mapping = {}
    for r in rows[1:]:
        code = r[code_idx]
        name = r[name_idx]
        if code is not None and name is not None:
            norm_name = normalize_name(str(name))
            mapping[norm_name] = {
                "codigo": code,
                "nome_original": name
            }
    print(f"✅ Mapeados {len(mapping)} códigos de empresas da planilha.")
    return mapping

def find_company_code(razao_social, mapping):
    """Busca o código interno da planilha a partir da Razão Social do SIGA."""
    siga_norm = normalize_name(razao_social)
    if siga_norm in mapping:
        return mapping[siga_norm]["codigo"], mapping[siga_norm]["nome_original"]
        
    for excel_norm, info in mapping.items():
        if len(siga_norm) > 4 and len(excel_norm) > 4:
            if siga_norm in excel_norm or excel_norm in siga_norm:
                return info["codigo"], info["nome_original"]
                
    return "", ""

def load_pfx_certs():
    """Escaneia a pasta raiz e a subpasta 'certificados' para encontrar os certificados .pfx."""
    pfx_files = []
    
    # Buscar na raiz do projeto
    for f in os.listdir("."):
        if f.endswith(".pfx"):
            pfx_files.append((f, f))
            
    # Buscar na subpasta 'certificados'
    pfx_dir = "certificados"
    if os.path.exists(pfx_dir) and os.path.isdir(pfx_dir):
        for f in os.listdir(pfx_dir):
            if f.endswith(".pfx"):
                pfx_files.append((os.path.join(pfx_dir, f), f))
                
    certs = {}
    for full_path, filename in pfx_files:
        match = re.search(r'\d{14}', filename)
        if match:
            cnpj = match.group(0)
            certs[cnpj] = full_path
    return certs

def load_pfx(pfx_path, password):
    """Carrega a chave e certificado a partir do arquivo .pfx."""
    with open(pfx_path, "rb") as f:
        pfx_data = f.read()
    private_key, certificate, _ = pkcs12.load_key_and_certificates(
        pfx_data, password.encode()
    )
    return private_key, certificate

def create_pem_files(private_key, certificate):
    """Gera arquivos PEM temporários em disco para o requests."""
    cert_pem = certificate.public_bytes(serialization.Encoding.PEM)
    key_pem = private_key.private_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PrivateFormat.TraditionalOpenSSL,
        encryption_algorithm=serialization.NoEncryption()
    )
    
    cert_fd, cert_path = tempfile.mkstemp(suffix=".pem")
    with os.fdopen(cert_fd, "wb") as f:
        f.write(cert_pem)
        
    key_fd, key_path = tempfile.mkstemp(suffix=".pem")
    with os.fdopen(key_fd, "wb") as f:
        f.write(key_pem)
        
    return cert_path, key_path

def download_xml_from_sefaz_by_key(cert_path, key_path, cnpj, key_nfe):
    """Chama o WebService SOAP da SEFAZ Nacional para baixar o XML da NF-e por chave de acesso."""
    wait_for_rate_limit()
    
    soap_body = (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<soap12:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:soap12="http://www.w3.org/2003/05/soap-envelope">'
        '<soap12:Body>'
        '<nfeDistDFeInteresse xmlns="http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe">'
        '<nfeDadosMsg>'
        f'<distDFeInt xmlns="http://www.portalfiscal.inf.br/nfe" versao="1.01">'
        f'<tpAmb>1</tpAmb>'
        f'<cUFAutor>23</cUFAutor>'  # 23 = Ceará
        f'<CNPJ>{cnpj}</CNPJ>'
        f'<consChNFe>'
        f'<chNFe>{key_nfe}</chNFe>'
        f'</consChNFe>'
        f'</distDFeInt>'
        '</nfeDadosMsg>'
        '</nfeDistDFeInteresse>'
        '</soap12:Body>'
        '</soap12:Envelope>'
    )
    
    url = "https://www1.nfe.fazenda.gov.br/NFeDistribuicaoDFe/NFeDistribuicaoDFe.asmx"
    headers = {
        'Content-Type': 'application/soap+xml; charset=utf-8; action="http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe/nfeDistDFeInteresse"',
        'User-Agent': 'Mozilla/5.0'
    }
    
    try:
        r = requests.post(url, headers=headers, data=soap_body, cert=(cert_path, key_path), verify=True, timeout=20)
        if r.status_code == 200:
            root = ET.fromstring(r.text)
            c_stat_el = root.find('.//{http://www.portalfiscal.inf.br/nfe}cStat')
            x_motivo_el = root.find('.//{http://www.portalfiscal.inf.br/nfe}xMotivo')
            c_stat = c_stat_el.text if c_stat_el is not None else ""
            x_motivo = x_motivo_el.text if x_motivo_el is not None else "Desconhecido"
            
            if c_stat == "138":
                doc_zip_el = root.find('.//{http://www.portalfiscal.inf.br/nfe}docZip')
                if doc_zip_el is not None and doc_zip_el.text:
                    raw_bytes = base64.b64decode(doc_zip_el.text)
                    xml_content = gzip.decompress(raw_bytes).decode('utf-8')
                    return True, xml_content, "OK"
                return False, None, "Erro: docZip vazio"
            elif c_stat == "137":
                return False, None, "Nota não cadastrada no Ambiente Nacional"
            else:
                return False, None, f"Rejeição SEFAZ ({c_stat}): {x_motivo}"
        return False, None, f"Erro HTTP {r.status_code}"
    except Exception as e:
        return False, None, f"Erro na requisição: {e}"

def extract_key_from_xml(xml_content):
    """Extrai a chave de acesso de 44 dígitos de dentro do XML descompactado."""
    match = re.search(r'<chNFe>(\d{44})</chNFe>', xml_content)
    if match:
        return match.group(1)
    match = re.search(r'Id="NFe(\d{44})"', xml_content)
    if match:
        return match.group(1)
    return None

def download_xmls_by_nsu(cert_path, key_path, cnpj, folder_path):
    """Baixa em lote todos os XMLs de notas destinadas por NSU (Sem limites por hora)."""
    url = "https://www1.nfe.fazenda.gov.br/NFeDistribuicaoDFe/NFeDistribuicaoDFe.asmx"
    headers = {
        'Content-Type': 'application/soap+xml; charset=utf-8; action="http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe/nfeDistDFeInteresse"',
        'User-Agent': 'Mozilla/5.0'
    }
    
    ult_nsu = "000000000000000"
    
    # Tentativa de carregar o último NSU salvo localmente para este CNPJ e evitar reprocessar tudo
    nsu_cache_file = os.path.join(folder_path, "ultimo_nsu.txt")
    if os.path.exists(nsu_cache_file):
        try:
            with open(nsu_cache_file, "r") as f_nsu:
                ult_nsu = f_nsu.read().strip()
                # Garante que seja preenchido com zeros à esquerda
                ult_nsu = f"{int(ult_nsu):015d}"
        except:
            ult_nsu = "000000000000000"
            
    print(f"  ➜ Iniciando busca por NSU a partir do NSU: {ult_nsu}")
    
    downloaded_nsu_count = 0
    while True:
        soap_body = (
            '<?xml version="1.0" encoding="utf-8"?>'
            '<soap12:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:soap12="http://www.w3.org/2003/05/soap-envelope">'
            '<soap12:Body>'
            '<nfeDistDFeInteresse xmlns="http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe">'
            '<nfeDadosMsg>'
            f'<distDFeInt xmlns="http://www.portalfiscal.inf.br/nfe" versao="1.01">'
            f'<tpAmb>1</tpAmb>'
            f'<cUFAutor>23</cUFAutor>'
            f'<CNPJ>{cnpj}</CNPJ>'
            f'<distNSU>'
            f'<ultNSU>{ult_nsu}</ultNSU>'
            f'</distNSU>'
            f'</distDFeInt>'
            '</nfeDadosMsg>'
            '</nfeDistDFeInteresse>'
            '</soap12:Body>'
            '</soap12:Envelope>'
        )
        
        try:
            r = requests.post(url, headers=headers, data=soap_body, cert=(cert_path, key_path), verify=True, timeout=30)
            if r.status_code != 200:
                print(f"    ❌ Erro de conexão com a SEFAZ (HTTP {r.status_code}). Parando loop.")
                break
                
            root = ET.fromstring(r.text)
            c_stat_el = root.find('.//{http://www.portalfiscal.inf.br/nfe}cStat')
            c_stat = c_stat_el.text if c_stat_el is not None else ""
            
            # Atualiza o NSU retornado pela SEFAZ
            ult_nsu_el = root.find('.//{http://www.portalfiscal.inf.br/nfe}ultNSU')
            max_nsu_el = root.find('.//{http://www.portalfiscal.inf.br/nfe}maxNSU')
            
            ret_ult_nsu = ult_nsu_el.text if ult_nsu_el is not None else ult_nsu
            ret_max_nsu = max_nsu_el.text if max_nsu_el is not None else "0"
            
            if c_stat == "138": # Documentos localizados
                doc_zips = root.findall('.//{http://www.portalfiscal.inf.br/nfe}docZip')
                for doc in doc_zips:
                    raw_bytes = base64.b64decode(doc.text)
                    xml_content = gzip.decompress(raw_bytes).decode('utf-8')
                    
                    # Extrai chave
                    key = extract_key_from_xml(xml_content)
                    if key:
                        xml_filename = os.path.join(folder_path, f"{key}.xml")
                        if not os.path.exists(xml_filename):
                            # Salva apenas se for um XML de nota completo (nfeProc) ou resumo útil (resNFe)
                            if "nfeProc" in xml_content or "resNFe" in xml_content:
                                with open(xml_filename, "w", encoding="utf-8") as f_xml:
                                    f_xml.write(xml_content)
                                downloaded_nsu_count += 1
                                
                ult_nsu = f"{int(ret_ult_nsu):015d}"
                # Salva o progresso do NSU localmente
                with open(nsu_cache_file, "w") as f_nsu:
                    f_nsu.write(ult_nsu)
                    
                # Se alcançou o máximo disponível na SEFAZ
                if int(ret_ult_nsu) >= int(ret_max_nsu):
                    break
            elif c_stat == "137": # Nenhum documento novo localizado
                break
            elif c_stat == "656": # Uso indevido (deve aguardar tempo padrão)
                print("    ⚠️ Rejeição 656: Uso indevido no método NSU. Aguardando 3 segundos...")
                time.sleep(3)
            else:
                x_motivo_el = root.find('.//{http://www.portalfiscal.inf.br/nfe}xMotivo')
                x_motivo = x_motivo_el.text if x_motivo_el is not None else "Desconhecido"
                print(f"    ❌ Rejeição SEFAZ ({c_stat}): {x_motivo}")
                break
                
            time.sleep(1.2) # Intervalo seguro recomendado pela SEFAZ
        except Exception as e:
            print(f"    ❌ Erro ao processar lote NSU: {e}")
            break
            
    print(f"  ➜ Consulta por NSU finalizada. XMLs baixados: {downloaded_nsu_count}")
    return downloaded_nsu_count

def extract_keys_from_xlsx(filepath):
    """Abre a planilha Excel e lê todas as chaves de acesso válidas (coluna 'Chave NF-e')."""
    if not os.path.exists(filepath):
        return []
        
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active
        headers = [str(cell.value).strip().lower() for cell in sheet[1]]
        
        chave_idx = -1
        for idx, h in enumerate(headers):
            if "chave" in h:
                chave_idx = idx
                break
                
        if chave_idx == -1:
            return []
            
        keys = []
        for row in list(sheet.iter_rows(min_row=2, values_only=True)):
            val = row[chave_idx]
            if val:
                key_clean = str(val).replace(" ", "").strip()
                if len(key_clean) == 44 and key_clean.isdigit():
                    keys.append(key_clean)
        return keys
    except Exception as e:
        print(f"⚠️ Erro ao abrir a planilha '{os.path.basename(filepath)}': {e}")
        return []

def main():
    print("=" * 60)
    print("   SIGA SEFAZ-CE - BAIXADOR DE XMLS DE NF-E VIA CERTIFICADO A1")
    print("=" * 60)
    
    # 1. Escanear certificados .pfx locais
    pfx_certs = load_pfx_certs()
    if not pfx_certs:
        print("❌ Nenhum certificado digital A1 (.pfx) encontrado na pasta raiz ou na pasta 'certificados'.")
        print("Adicione seus arquivos de certificado no formato 'NomeEmpresa_CNPJ.pfx' no diretório raiz ou dentro da pasta 'certificados'.")
        sys.exit(1)
        
    print(f"✅ Encontrados {len(pfx_certs)} certificados A1 (.pfx) mapeados para download.")
    
    # 2. Carregar Excel de-para
    excel_mapping = load_excel_mapping(EXCEL_FILE)
    
    # 3. Solicitar mês/ano
    try:
        ano = int(input("\nDigite o ano de referência (ex: 2026): ").strip())
        mes_num = int(input("Digite o número do mês de referência (1 a 12): ").strip())
        if mes_num < 1 or mes_num > 12:
            raise ValueError()
    except ValueError:
        print("❌ Ano ou mês inválido. Saindo.")
        sys.exit(1)
        
    date_subfolder = f"{mes_num:02d}{ano}"
    
    # 4. Solicitar senha padrão do certificado
    password = input("\nDigite a senha padrão dos certificados: ").strip()
    
    # 5. Processamento dos downloads
    processed_count = 0
    downloaded_count = 0
    failed_count = 0
    
    for cnpj_cert, pfx_filename in pfx_certs.items():
        company_dirs = [d for d in os.listdir(".") if os.path.isdir(d) and d.endswith("-")]
        
        target_dir = None
        for d in company_dirs:
            if cnpj_cert in d:
                target_dir = d
                break
                
        if not target_dir:
            razao_social = ""
            if os.path.exists("siga_empresas.csv"):
                try:
                    with open("siga_empresas.csv", mode="r", encoding="utf-8-sig") as f_csv:
                        reader = csv.DictReader(f_csv, delimiter=";")
                        for row in reader:
                            c_csv = "".join(filter(str.isdigit, row.get("cnpj", "")))
                            if cnpj_cert in c_csv:
                                razao_social = row.get("razaoSocial", "")
                                break
                except:
                    pass
            
            if razao_social:
                cod_excel, nome_excel = find_company_code(razao_social, excel_mapping)
                company_label = f"{cod_excel} - {nome_excel}" if cod_excel else f"{cnpj_cert} - {razao_social}"
                company_label = sanitize_filename(company_label)
                target_dir = f"{company_label}-"
            else:
                target_dir = f"{cnpj_cert}-"
                
        folder_path = os.path.join(target_dir, date_subfolder)
        if not os.path.exists(folder_path):
            print(f"\n📂 Pasta não encontrada para CNPJ {cnpj_cert}: '{folder_path}' (Pulando)")
            continue
            
        print(f"\n📁 Lendo pasta: {target_dir} ({cnpj_cert})")
        
        # Carregar o certificado da empresa em memória
        cert_loaded = False
        cert_path = None
        key_path = None
        
        # Tenta com a senha informada
        pwd_tentar = password
        for attempt in range(3):
            try:
                private_key, certificate = load_pfx(pfx_filename, pwd_tentar)
                cert_path, key_path = create_pem_files(private_key, certificate)
                cert_loaded = True
                break
            except Exception:
                print(f"  ❌ Senha inválida para o certificado '{pfx_filename}'!")
                if attempt < 2:
                    pwd_tentar = input(f"  Digite a senha correta para '{pfx_filename}': ").strip()
                else:
                    print(f"  ❌ Falha ao carregar o certificado '{pfx_filename}'. Pulando empresa.")
                    
        if not cert_loaded:
            continue
            
        # ----------------------------------------------------
        # FLUXO 1: Destinatário (Entradas) - Via NSU (Sem Limite)
        # ----------------------------------------------------
        print("  [Processando Entradas/Destinatário via NSU]")
        downloaded_count += download_xmls_by_nsu(cert_path, key_path, cnpj_cert, folder_path)
        
        # ----------------------------------------------------
        # FLUXO 2: Emissor (Saídas) - Via Chave (Limite 20/hora)
        # ----------------------------------------------------
        print("  [Processando Saídas/Emissor via Chave]")
        xlsx_emissor = os.path.join(folder_path, f"nfe_emissor_{date_subfolder}.xlsx")
        keys_emissor = extract_keys_from_xlsx(xlsx_emissor)
        
        if keys_emissor:
            print(f"  ➜ Encontradas {len(keys_emissor)} chaves de saída na planilha.")
            for idx, key in enumerate(keys_emissor):
                xml_filename = os.path.join(folder_path, f"{key}.xml")
                if os.path.exists(xml_filename):
                    continue
                    
                processed_count += 1
                print(f"    [{idx+1}/{len(keys_emissor)}] Baixando Saída {key}... ", end="", flush=True)
                
                success, xml_content, msg = download_xml_from_sefaz_by_key(cert_path, key_path, cnpj_cert, key)
                
                if success:
                    with open(xml_filename, "w", encoding="utf-8") as f_xml:
                        f_xml.write(xml_content)
                    print("✅ Sucesso!")
                    downloaded_count += 1
                else:
                    print(f"❌ Falha: {msg}")
                    failed_count += 1
                    
                # Pequeno delay entre requisições
                time.sleep(0.5)
        else:
            print("  ➜ Nenhuma chave de saída na planilha de emissor.")
            
        # Garantir limpeza dos PEMs temporários do certificado atual
        if cert_path and os.path.exists(cert_path):
            os.unlink(cert_path)
        if key_path and os.path.exists(key_path):
            os.unlink(key_path)
                
    print("\n" + "=" * 60)
    print("🎉 EXECUÇÃO CONCLUÍDA COM SUCESSO!")
    print(f"   - Total de XMLs baixados: {downloaded_count}")
    print(f"   - Total de falhas em saídas: {failed_count}")
    print("=" * 60)

if __name__ == "__main__":
    main()
